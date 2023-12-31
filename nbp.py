from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import requests
import json
import datetime
import os


wb = Workbook()
ws = wb.active
ws.tittle = "kursy"

x = datetime.date.today()
y = x.month - 2
z = x.year - 1

if x.month == 1:
    y = 11

if x.month == 2:
    y = 12
q = datetime.date(z,y,x.day)
p = datetime.date(z,12,31)
print(q,p)

if y== 12 or y == 11:
    url = f"https://api.nbp.pl/api/exchangerates/tables/A/2023-01-01/{x}/?format=json"
    url2 = f"https://api.nbp.pl/api/exchangerates/tables/A/{q}/{z}-12-31/?format=json"
    response = requests.get(url)
    print(url)
    response2 = requests.get(url2)
    data = json.loads(response.text)
    data2 = json.loads(response2.text)

    ws["C1"] = data2[0]["no"]
    ws["B1"] = "waluta2"
    ws["A1"] = "waluta"
    ws.column_dimensions['C'].width = 15

    for r in range(0, len(data2[0]["rates"])):
        ws["A" + str(r + 2)] = str(data2[0]["rates"][r]["currency"])
        ws["B" + str(r + 2)] = str(data2[0]["rates"][r]["code"])
        ws["C" + str(r + 2)] = str(data2[0]["rates"][r]["mid"])

    for col in range(0, len(data2)):
        char = get_column_letter(col + 3)
        ws[char + str(1)] = data2[col]["no"]
        for row in range(0, len(data2[0]["rates"])):
            ws[char + str(row + 2)] = data2[col]["rates"][row]["mid"]
            ws.column_dimensions[char].width = 15
    col = col + 1

    for col1 in range(0, len(data)):
        char1 = get_column_letter(col1 + col + 3)
        ws[char1 + str(1)] = data[col1]["no"]
        for row in range(0, len(data[0]["rates"])):
            ws[char1 + str(row + 3)] = data[col1]["rates"][row]["mid"]
            ws.column_dimensions[char1].width = 15

    tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(col +col1 + 3)}{row+2}")

    style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    ws.add_table(tab)
else:

    print(len({x.day}))
    if len(str({x.day})) == 1:
        m = f"0{x.day}"
    else:
        m = f"{x.day}"

    if len(str({y})) == 1:
        y = f"0{y}"
    else:
        y = f"{y}"
    print(f"{x.year}-{y}-{m}")


    url3 = f"https://api.nbp.pl/api/exchangerates/tables/A/{x.year}-0{y}-{m}/{x}/?format=json"
    response3 = requests.get(url3)
    data3 = json.loads(response3.text)

    ws["C1"] = data3[0]["no"]
    ws["B1"] = "waluta2"
    ws["A1"] = "waluta"
    ws.column_dimensions['C'].width = 15

    for r in range(1, len(data3[0]["rates"])):
        ws["A" + str(r + 1)] = data3[0]["rates"][r]["currency"]
        ws["B" + str(r + 1)] = data3[0]["rates"][r]["code"]
        ws["C" + str(r + 1)] = data3[0]["rates"][r]["mid"]

    for col in range(1, len(data3)):
        char = get_column_letter(col + 3)
        ws[char + str(1)] = data3[col]["no"]
        for row in range(1, len(data3[0]["rates"])):
            ws[char + str(row + 1)] = data3[col]["rates"][row]["mid"]
            ws.column_dimensions[char].width = 15

    tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(col + 3)}{row+1}")

    style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    ws.add_table(tab)

wb.save('excel NBP.xlsx')

os.startfile('excel NBP.xlsx')